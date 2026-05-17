"""Export utilities — Excel and CSV export for any DataTable."""
import csv
import os
from datetime import datetime


def export_to_csv(headers: list, rows: list, filename: str = None) -> str:
    """Export rows to CSV, return file path."""
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{ts}.csv"
    path = os.path.join(os.path.expanduser("~"), "Downloads", filename)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    return path


def export_to_excel(headers: list, rows: list, sheet_name: str = "Sheet1",
                    filename: str = None) -> str:
    """Export rows to Excel .xlsx, return file path."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{ts}.xlsx"
    path = os.path.join(os.path.expanduser("~"), "Downloads", filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1A2035")
    hdr_font = Font(bold=True, color="F5A623", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="2A3347")
    border = Border(bottom=Side(style="medium", color="F5A623"))

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = max(len(str(h)) + 4, 14)

    # Data rows
    alt_fill = PatternFill("solid", fgColor="0A0F1E")
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=str(val) if val is not None else "")
            cell.alignment = Alignment(vertical="center")
            if r % 2 == 0:
                cell.fill = alt_fill

    ws.row_dimensions[1].height = 24
    wb.save(path)
    return path


def export_table_widget(table_widget, sheet_name="Export", filename=None, fmt="excel") -> str:
    """Export a QTableWidget directly."""
    headers = []
    for c in range(table_widget.columnCount()):
        h = table_widget.horizontalHeaderItem(c)
        headers.append(h.text() if h else f"Col{c+1}")

    rows = []
    for r in range(table_widget.rowCount()):
        row = []
        for c in range(table_widget.columnCount()):
            item = table_widget.item(r, c)
            row.append(item.text() if item else "")
        rows.append(row)

    if fmt == "csv":
        return export_to_csv(headers, rows, filename)
    return export_to_excel(headers, rows, sheet_name, filename)


def open_file(path: str):
    """Open file with the OS default application."""
    import subprocess, platform
    system = platform.system()
    if system == "Windows":
        os.startfile(path)
    elif system == "Darwin":
        subprocess.run(["open", path])
    else:
        subprocess.run(["xdg-open", path])
